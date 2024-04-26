import re
import os
import time
import shutil
import openpyxl
import datetime
import threading
import telebot
import sqlite3
import requests
from telebot import types


# Папки для работы системы
otkyda = os.path.join(os.path.dirname(__file__), "Downloads/")
kyda = os.path.join(os.path.dirname(__file__), 'pyp/')
# Количество сообщений, необходимое для спам-атаки
SPAM_LIMIT = 8
# Словарь для хранения количества сообщений, отправленных каждым пользователем
spam_count = {}
# Текст сообщения, отправляемого пользователю
WARNING_MESSAGE = "Слишком много сообщений за короткое время. Попробуйте позже."
# Время, в течение которого считается количество сообщений пользователя
SPAM_TIME_LIMIT = 35
# Админы, есть некоторые плюшки
admin = [650093184, 650093184]
# Время обновления .xlsx
im_done = 0
noww = datetime.datetime.now()
reboot_time = noww.strftime("%H:%M:%S")
say = ""
stroka = ""  # type: ignore
zachem_message = '❓Зачем вы меняете класс?\n😅Возможно вы хотите узнать расписание друзей на завтра, тогда просто напишите их класс и плюсик в конце, например так: "11Б+" (без ковычек)\n✍🏻Если причина не в этом, пожалуйста напишите мне @maryin_n, постараюсь помочь'

# объявляем бота
bot = telebot.TeleBot('6417160051:AAE4yZDMbXw91trVBLSWASmGh0fxxxN-_U0')  # релизный бот

# Определение учителей и классов, которые они ведут

teacher = ['НИНАНИКОЛАЕВНА', 'АНАСТАСИЯРОМАНОВНА', 'ГАЛИНАМИХАЙЛОВНА', 'АНАСТАСИЯСЕРГЕЕВНА']
name_teacher = ['Нина Николаевна', 'Анастасия Романовна', 'Галина Михайловна', 'Анастасия Сергеевна']
teacher_urok = ["История", "нформат", "География", "Англ"]
teacher_urok_for_zam = [["История"], ["Информатика", "Основы информатики"], ["География"], ["Англ.яз"]]
teacher_klass = [["8Г", "8Д", "10А", "10Б", "10В", "10Г", "10Д", "11А", "11Б", "11В", "11Г", "11Д"],
                 ["5А", "5Б", "5В", "5Г", "5Д", "5Е", "6А", "6Б", "6В", "6Г", "7А", "7Б", "7В", "7Г", "11Б", "11Г"],
                 ['5А', '5Б', '5В', '5Г', '5Д', '5Е'], ['5А', '5Е', "9А", "11А", "11Б", "11В", "11Г", "11Д"]]
klass_ruk = ["", "11Б", "", "11Д"]

# Списки существующих классов

stok_klass = ['5-', '6-', '7-', '8-', '9-', '10-', '11-',
              '5А', '5Б', '5В', '5Г', '5Д', '5Е', '5A', '5B', '5V', '5G', '5D', '5E', '6А', '6Б', '6В', '6Г', '6Д',
              '6Е', '6A',
              '6B', '6V', '6G', '6D', '6E', '7А', '7Б', '7В', '7Г', '7Д', '7Е', '7A', '7B', '7V', '7G', '7D', '7E',
              '8А', '8Б',
              '8В', '8Г', '8Д', '8Е', '8A', '8B', '8V', '8G', '8D', '8E', '9А', '9Б', '9В', '9Г', '9Д', '9Е', '9A',
              '9B', '9V',
              '9G', '9D', '9E', '10А', '10Б', '10В', '10Г', '10Д', '10Е', '10A', '10B', '10V', '10G', '10D', '10E',
              '11А',
              '11Б', '11В', '11Г', '11Д', '11Е', '11A', '11B', '11V', '11G', '11D', '11E']

# Преобразование списков классов для нормальной работы

awe = ["/" + i for i in stok_klass + teacher]

qwq = [i + "+" for i in stok_klass + teacher] + stok_klass + teacher

# это директория базы данных
os.chdir(kyda)

std_keyboard = types.ReplyKeyboardMarkup(True, True)  # Определение стартовой клавиатуры
std_keyboard.row("На сегодня")
std_keyboard.row("Настройки", "На завтра")  # Задаём кнопки : Настройки, На сегодня, На завтра

# коннект к базе данных
conn = sqlite3.connect('userdate.db', check_same_thread=False)
cursor = conn.cursor()
# создаем базу данных, если она еще не существует
cursor.execute('''CREATE TABLE IF NOT EXISTS userdata
             (id INTEGER PRIMARY KEY AUTOINCREMENT,
             usertelegramid TEXT NOT NULL,
             class TEXT NOT NULL, settings TEXT DEFAULT '00000000', active INTEGER DEFAULT 0, user_tag TEXT)''')
# обновляем базу
conn.commit()

logf = open('log.txt', 'a', encoding='utf-8')


def log(func):
    def wrapper(*args, **kwargs):
        try:
            arg = args
            start_time = datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')
            result = func(*args, **kwargs)
            end_time = datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')
            logf.write(f'[{start_time}] {func.__name__} started with {arg}, {kwargs}\n')
            logf.write(f'[{end_time}] {func.__name__} finished with {result}\n')
            return result
        except Exception as e:
            start_time = datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')
            logf.write(f'[{start_time}] ERROR at {func.__name__}: {str(e)}\n')
            raise e

    return wrapper


def clear_log():
    logf.truncate(0)
    start_time = datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')
    logf.write(f'[{start_time}] BOT HAS BEEN RESTARTED\n')
    logf.write(f'[{start_time}] LOG CLEARED\n')


def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value


# получение настроек пользователя
def get_settings_from_id(id):
    cursor.execute(f"SELECT settings FROM userdata WHERE usertelegramid=?", (id,))
    settings = cursor.fetchall()
    return settings


def holyday(x):
    days = ["01.09", "08.03", "05.10", "23.01", "09.04", "01.04", "12.03", "29.12", "30.12", "31.12"]
    text = ["🔔С первым сентября!", "💐С 8 марта!", "🧑‍🏫С днём учителя наших педагогов!", "🛡С 23 февраля!", "🎗️С 9 мая!",
            "🌷С 1 мая!", "🚀С днём космонавтики!", "🎄С наступающим Новым годом!", "🎄С наступающим Новым годом!",
            "🎄С наступающим Новым годом!"]
    year = int(datetime.date.today().strftime('%y'))
    for i in range(len(days)):
        if x == days[i]: return text[i]
    if x == '12.09' or '13.09':
        if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0:
            if x == '12.09': return "С++ днём програмиста!"
        else:
            if x == '13.09': return "С++ днём програмиста!"
    return 0


# функция формирующая список который обрабатывает wiwod
@log
def xxx(need_class, sz, id=1, one=0, name="", klass="qwerty01"):
    os.chdir(kyda)
    skip = 1
    if one:
        current_file = openpyxl.load_workbook(filename="1.xlsx")
    else:
        if name:
            try:
                current_file = openpyxl.load_workbook(
                    filename=f"{kyda}{name}.xlsx")
            except:
                print("err opn")
                return "Это сообщение заглушка. Извините, у нас произошла ошибка."
        elif sz == 1:
            try:
                current_file = openpyxl.load_workbook(
                    filename=f"{kyda}{(datetime.date.today()).strftime('%d.%m.%Y')}.xlsx")

                # current_file = openpyxl.load_workbook(filename="C:\\pyp\\09.10.2023.xlsx")
            except:
                return "Расписаниe на сегодня отсутствует"
        elif sz == 2:
            try:
                skip = 1
                if int(datetime.datetime.today().weekday()) == 4:
                    skip = 3
                elif int(datetime.datetime.today().weekday()) == 5:
                    skip = 2
                current_file = openpyxl.load_workbook(
                    filename=f"{kyda}{(datetime.date.today() + datetime.timedelta(days=skip)).strftime('%d.%m.%Y')}.xlsx")
            except:
                if str(id) in (list(subscribers().split())):
                    return "Расписания на завтра пока нет"
                else:
                    return "Расписания на завтра пока нет.\nНо вы можете подписаться на автоматическую рассылку расписания написав /subscribe"
    now_have = current_file.sheetnames
    table = current_file[now_have[0]]
    classes = []
    times = []
    cell = None  # type: ignore
    if sz == 1:
        x = datetime.date.today().strftime('%d.%m')
    else:
        x = (datetime.date.today() + datetime.timedelta(days=skip)).strftime('%d.%m')
    holydayli = holyday(x)
    if holydayli and (int(str((get_settings_from_id(id))[0][0])[3]) == 0):
        st = holydayli + "\n"
    else:
        st = ""
    if st:
        classes.append("")
        times.append(holydayli)
    classes.append("")
    nakakoe = table.cell(row=1, column=2).value  # type: ignore
    if nakakoe:
        if "уроков" in nakakoe:
            nakakoe = nakakoe[:nakakoe.find("уроков")] + nakakoe[nakakoe.find("уроков") + 7:]
        if "Изменения расписания" in nakakoe:
            nakakoe = nakakoe.replace("Изменения расписания", "Расписание")
        times.append("🗓" + nakakoe)
        data = "🗓" + nakakoe
    else:
        times.append("🗓" + f"Расписание на {x}.2023")
        data = "🗓" + f"Расписание на {x}.2023"

    time_cell_coord = 2
    while re.fullmatch(r'\d{2}:\d{2} - \d{2}:\d{2}', table.cell(row=10, column=time_cell_coord).value) is None:
        time_cell_coord += 1

    def get_shedule(need_class):
        number = 1
        for row in table.iter_rows():  # type: ignore
            for cell in row:
                if cell.value == need_class:
                    row_num = cell.row
                    col_num = cell.column
                    for i in range(row_num + 1, row_num + 12):

                        if re.fullmatch(r'\d{2}:\d{2} - \d{2}:\d{2}',
                                        str(table.cell(row=i,
                                                       column=time_cell_coord).value)) is not None:  # type: ignore
                            cell = table.cell(row=i, column=col_num)  # type: ignore
                            if isinstance(cell, openpyxl.cell.cell.MergedCell):  # type: ignore
                                classes.append(getMergedCellVal(table, cell))

                            # ЙА ПАДЖИГАТЕЛЬ, ПАДЖИГАТЕЛЬ

                            elif re.fullmatch(r'=\$\w+\$\d+', str(cell.value)) is not None:
                                s = cell.value.replace('=', '').replace('$', '')
                                classes.append(table[s].value)
                            # БОЛЬШЕ НЕ ПАДЖИГАТЕЛЬ :(

                            else:
                                classes.append(cell.value)  # type: ignore
                            cell = table.cell(row=i, column=time_cell_coord)  # type: ignore
                            if int(str((get_settings_from_id(id))[0][0])[4]) == 1 and not (klass in teacher):
                                times.append(f"({number}) {cell.value}")
                                number += 1
                            else:
                                times.append(cell.value)  # type: ignore
                        else:
                            break
        return list(map(lambda x, y: [x, y], times, classes))

    def les_check(s: str) -> re.Match:
        return re.fullmatch(
            r'[-a-zA-ZА-Яа-яёЁ_.:,]+((\d+)|(Геол.музей)|([БМ].спорт.зал)|(Ист.музей)|(Акт.зал))((/[-a-zA-ZА-Яа-яёЁ_.:,]+\d+)|(/\d+)|(/[-a-zA-ZА-Яа-яёЁ_.:,]+((\d+)|(Геол.музей)|([БМ].спорт.зал)|(Ист.музей)|(Акт.зал))))*',
            s)

    def simplify_string(st, f):
        if f not in st:
            raise IndexError(f"In func: simplify_string - {f} not in {st}")
        if f.isdigit() and f not in re.findall(r'\d+', st):
            raise IndexError(f"In func: simplify_string - {f} not in {st}")
        for i in ['Геол.музей', "Б.спорт.зал", "М.спорт.зал", "Ист.музей",
                  "Акт.зал"]:  # приводим название кабинетов к нижнему регистру (нужно для работы далее)
            if i in st:
                st = re.sub(i, i.lower(), st)

        f = f.lower()
        st = st.replace(' ', '').replace('\n', '')

        if re.search(r'/[A-ZА-ЯЁ]', st):

            x = re.search(r'/[A-ZА-ЯЁ]', st).start()  # Разделяем строку на две по обр. слэшу и большой букве
            a = st[:x]
            b = st[x + 1:]

            if f.isdigit():
                if f in re.findall(r'\d+', a):
                    return re.match(r'[-A-ZА-ЯЁ_:,.][-a-zа-яё_:,.]+', a[:a.index(f)]).group()
                else:
                    return re.match(r'[-A-ZА-ЯЁ_:,.][-a-zа-яё_:,.]+', b[:b.index(f)]).group()

            elif f in ['геол.музей', "б.спорт.зал", "м.спорт.зал", "ист.музей", "акт.зал"]:
                if f in a:
                    return re.match(r'[-A-ZА-ЯЁ_:,.][-a-zа-яё_:,.]+', a[:a.index(f)]).group()
                else:
                    return re.match(r'[-A-ZА-ЯЁ_:,.][-a-zа-яё_:,.]+', b[:b.index(f)]).group()

        else:
            return re.match(r'[-A-ZА-ЯЁ_:,.][-a-zа-яё_:,.]+', st[:st.index(f)]).group()

    if need_class.isdigit() or need_class in ['Геол.музей', "Б.спорт.зал", "М.спорт.зал", "Ист.музей", "Акт.зал"]:
        time = []
        klass = []
        # t_r = table.iter_rows()
        for row in table.iter_rows():
            for cell in row:
                if cell.value:
                    s = str(cell.value).replace('\n', '').replace(' ', '')
                    if les_check(s):
                        try:
                            if (simplify_string(s, need_class)):
                                x = ''
                                for i in range(cell.row, max(cell.row - 12, 0), -1):
                                    if table.cell(row=i, column=cell.column).value:
                                        if re.fullmatch(r'\d{1,2}[A-ZА-Я]',
                                                        table.cell(row=i, column=cell.column).value):
                                            x = table.cell(row=i, column=cell.column).value
                                            break
                                klass.append(f'{simplify_string(s, need_class)} {x}')
                                time.append(str(table.cell(row=cell.row, column=time_cell_coord).value))
                        except:
                            continue
        return sorted(list(map(lambda x, y: [x, y], time, klass)), key=lambda x: x[0])


    def standart_check(a: list[list], klass, yrok, klass_ruk_to):
        times_standart = []
        classes_standart = []
        for i in a:
            if klass_ruk_to != "":
                if (i[-1] is not None and yrok in i[-1]) or (i[-1] is not None and (
                        'Разговоры о важном' in i[-1] or 'Россия -' in i[-1]) and klass == klass_ruk_to):
                    times_standart.append(i[0])
                    classes_standart.append(i[1])
            else:
                if (i[-1] is not None and yrok in i[-1]):
                    times_standart.append(i[0])
                    classes_standart.append(i[1])
        # print(list(map(lambda x, y: [x, y], times_standart, classes_standart)))
        return list(map(lambda x, y: [x, y], times_standart, classes_standart))



    if need_class in teacher:
        for j in range(len(teacher)):
            if need_class == teacher[j]:
                r = teacher_klass[j]
                s = []
                x = []
                for i in r:

                    to_append = standart_check(xxx(i, sz, id, 0, "", need_class), i, teacher_urok[j], klass_ruk[j])
                    if len(to_append) == 1:
                        if teacher_urok[j] == "Англ":
                            if "\n" in str(to_append[0][1]):
                                to_append = [[to_append[0][0], str(to_append[0][1])[:str(to_append[0][1]).find("\n")]]]
                                if str(to_append[0][1][-1]) == "/": to_append = [
                                    [to_append[0][0], str(to_append[0][1])[:-1]]]
                        s.append(to_append)  # type: ignore
                    elif len(to_append) == 0:
                        continue
                    else:
                        for t in range(len(to_append)):
                            if teacher_urok[j] == "Англ":
                                if "\n" in str(to_append[t][1]):
                                    to_append = [
                                        [to_append[j][0], str(to_append[j][1])[:str(to_append[j][1]).find("\n")]]]
                                    if str(to_append[j][1][-1]) == "/": to_append = [
                                        [to_append[j][0], str(to_append[j][1])[:-1]]]
                            s.append([to_append[t]])

                    if len(to_append) == 1:
                        try:
                            for iii in range(5):
                                s[-1][iii][1] = s[-1][iii][1] + f" 📍 {i}"
                        except:
                            continue
                    else:
                        for g in range(1, len(to_append) + 1):
                            try:
                                for iii in range(5):
                                    s[-g][iii][1] = s[-g][iii][1] + f" 📍 {i}"
                            except:
                                continue
                for i in s:
                    if i:
                        x.append(i[0])

                y = str(x)
                if not ("08:00 - 08:40" in y): x.append(["08:00 - 08:40", ""])
                if not ("08:50 - 09:30" in y) and (
                        "08:50 - 09:30" in y or "09:50 - 10:30" in y or "10:50 - 11:30" in y or "11:50 - 12:30" in y or "12:40 - 13:20" in y or "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["08:50 - 09:30", ""])
                if not ("09:50 - 10:30" in y) and (
                        "09:50 - 10:30" in y or "10:50 - 11:30" in y or "11:50 - 12:30" in y or "12:40 - 13:20" in y or "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["09:50 - 10:30", ""])
                if not ("10:50 - 11:30" in y) and (
                        "10:50 - 11:30" in y or "11:50 - 12:30" in y or "12:40 - 13:20" in y or "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["10:50 - 11:30", ""])
                if not ("11:50 - 12:30" in y) and (
                        "11:50 - 12:30" in y or "12:40 - 13:20" in y or "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["11:50 - 12:30", ""])
                if not ("12:40 - 13:20" in y) and (
                        "12:40 - 13:20" in y or "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["12:40 - 13:20", ""])
                if not ("13:30 - 14:10" in y) and (
                        "13:30 - 14:10" in y or "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["13:30 - 14:10", ""])
                if not ("14:30 - 15:10" in y) and (
                        "14:30 - 15:10" in y or "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["14:30 - 15:10", ""])
                if not ("15:30 - 16:10" in y) and (
                        "15:30 - 16:10" in y or "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["15:30 - 16:10", ""])
                if not ("16:30 - 17:10" in y) and (
                        "16:30 - 17:10" in y or "17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["16:30 - 17:10", ""])
                if not ("17:20 - 18:00" in y) and ("17:20 - 18:00" in y or "18:05 - 18:45" in y): x.append(
                    ["17:20 - 18:00", ""])
                if not ("18:05 - 18:45" in y) and ("18:50 - 19:30" in y): x.append(["18:05 - 18:45", ""])
                x.append(["Конец", ""])
            else:
                continue
        x = sorted(x, key=lambda i: i[0])
        if data:
            x = [[data, ""]] + x
        if holydayli:
            x = [[holydayli, ""]] + x
        return x
    else:
        return get_shedule(need_class)

# функция получения времени
def time123():
    now = datetime.datetime.now()
    timeprow(now.strftime("%H:%M:%S"))


# функция которая нужна для работы wiwid2
def forwiw(s):
    mx = len(s)
    c = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "Геол.музей", "Б.спорт.зал", "М.спорт.зал", "акт.зал"]
    for i in c:
        if i in s:
            if s.find(i) < mx:
                mx = s.find(i)
    return mx

# функция вывода строки расписания
@log
def wiwod2(h, id, sz, klass):
    sett = get_settings_from_id(id)
    if h == "Расписания на завтра пока нет.\nНо вы можете подписаться на автоматическую рассылку расписания написав /subscribe":
        return "Расписания на завтра пока нет.\nНо вы можете подписаться на автоматическую рассылку расписания написав /subscribe"
    if h == "Расписания на завтра пока нет":
        return "Расписания на завтра пока нет"
    if h == "Расписаниe на сегодня отсутствует":
        return "Расписаниe на сегодня отсутствует"
    data = ""
    st = ""
    if not (klass.isdigit() or klass in ['Геол.музей', 'Б.спорт.зал', 'Ист.музей']):
        if int(str(sett[0][0])[5]) == 1:
            if not (klass in teacher):
                data = h[0]
                h = h[1:]
                h = [[i[0][:i[0].rindex(" ")], i[1]] for i in h]
    else:
        st = f"📌Расписание уроков в кабинете {klass}\n"
    if int(str(sett[0][0])[2]) == 1:
        if len(klass) < 5:
            st = f"📌Для {klass} класса\n"
        else:
            st = f"📌Для учителя: {name_teacher[teacher.index(klass)]}\n"
    if int(str(sett[0][0])[1]) == 1:
        h = h[1:]
    if int(str(sett[0][0])[0]) == 0:
        gag = 1
    else:
        gag = 0

    chto = ["гебраических зада", "ометричиских зада", "физичес", "Всеобщая история", "изическая культу",
            "ометрическ", "Изобразительное", "Родная литература", "Физика в задачах", "Родной язык",
            "История России", "Всеобщая история", "Биология в задачах", "альный проект", "именение математических",
            "оссия - мои", "WEB", "Основы информатики", "Вероятность", "иологических задач", "решению физ задач",
            "Основы духовно-нравственной", "химических задач", "Основы программирования"]
    nachto = ["Математика", "Геометрия", "Физика", "История", "Физ-ра", "Геометрия", "ИЗО", "Литература", "Физика",
              "Русский язык", "История", "История", "Биология с экспериментами", "Индив. проект", "Математика",
              "РМГ", "WEB-разраб. и бд", "Информатика", "Вер. и стат.", "Биология", "Физика", "ОДКНР", "Химия",
              "Информатика"]
    c = ["4", "3", "2", "1", "5", "6", "7", "8", "9", "0", "Геол.музей", "Б.спорт.зал", "М.спорт.зал", "акт.зал",
         "Ист.музей"]
    for i in range(len(h)):
        st += str((h[i])[0])
        strok = str((h[i])[1])

        if strok != "None":
            flag = True
            for dd in c:
                if dd in strok and strok:
                    flag = False
                    break
            if flag:
                st += " " + strok + "\n"

        if strok == "None":
            st = st + " None" + "\n"
        if "Английский язык" in strok:
            strok = strok[:strok.find("Английский язык")] + "Англ.яз" + strok[(strok.find("Английский язык") + 15):]
            if "\n" in strok:
                strok = strok[:strok.find("\n")] + strok[(strok.find("\n") + 1):]
        if "Немецкий язык" in strok:
            strok = strok[:strok.find("Немецкий язык")] + "Нем.яз " + strok[(strok.find("Немецкий язык") + 13):]
            if "\n" in strok:
                strok = strok[:strok.find("\n")] + strok[(strok.find("\n") + 2):]
        if forwiw(strok):
            for j in c:
                if j in str(strok):
                    if not (klass.isdigit() or klass in ['Геол.музей', 'Б.спорт.зал', 'Ист.музей']):
                        st += " "
                        est = str(strok)[:forwiw(strok)]
                        for t in range(len(chto)):
                            if chto[t] in est and gag:
                                est = nachto[t]
                        kab = str(strok)[forwiw(strok):]
                        st += est + " " + kab
                        st += "\n"
                        break
                    else:
                        st += " "
                        est = str(strok)[:forwiw(strok)]
                        for t in range(len(chto)):
                            if chto[t] in est and gag:
                                est = nachto[t]
                        kab = str(strok)[forwiw(strok):]
                        st += est + " || " + kab
                        st += "\n"
                        break
    st = st.replace("  ", " ")
    if not (klass.isdigit() or klass in ['Геол.музей', 'Б.спорт.зал', 'Ист.музей']):
        if int(str((sett)[0][0])[4]) == 1:
            pass  # st = ciferki(st, klass)
        if int(str(sett[0][0])[5]) == 1 and not (klass in teacher):
            if data:
                st = str(data[0]) + "\n" + st.replace("- ", "")
            st = st.replace("- ", "")
        if int(str(sett[0][0])[6]) == 1 and klass in teacher:
            zamens = teacher_urok_for_zam[teacher.index(klass)]
            for i in range(len(zamens)):
                st = st.replace(zamens[i], "каб.")
    else:
        pod_zam = ["/" + str(lkj) for lkj in range(3, 60)] + [str(jkl) + "/" for jkl in range(3, 60)]
        for zam in pod_zam:
            st = st.replace(f"{zam}", " ")
            st = st.replace("  ", " ")
        st = st.replace("|| ||", "||")

    if gag:
        st = st.replace("Разговоры о важном", "Разг. о важн.")
    if "=" in st:
        print("=")
        st = st + "\n❗️❗️❗️У нас ошибка, посмотрите вручную, запросить таблицу за сегодня можно через /givemetoday, на завтра /givemetomorrow❗️❗️❗️"
        st = st.replace("=", "err")
        st = st.replace("$", "err")
    return st


def reboot():
    clear_log()
    os.system("sudo reboot")

def planreboot():
    time.sleep(14500)
    reboot()


# функция считает пользователей
@log
def count_users():
    conn = sqlite3.connect('userdate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM userdata WHERE usertelegramid > 0")
    count = cursor.fetchone()[0]
    conn.close()
    return count


# функция считает группы
@log
def count_grupp():
    conn = sqlite3.connect('userdate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM userdata WHERE usertelegramid < 0")
    count = cursor.fetchone()[0]
    conn.close()
    return count


# функция считает active пользователей
@log
def active_users():
    conn = sqlite3.connect('userdate.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM userdata WHERE active > 0")
    # cursor.execute(f"UPDATE userdata SET active=1 WHERE usertelegramid=?", (id,))
    count = cursor.fetchone()[0]
    conn.close()
    return count


# Define a function to delete a user
def delete_user(user_id):
    cursor.execute("DELETE FROM userdata WHERE usertelegramid=?", (user_id,))
    conn.commit()
    print("User deleted successfully")


def zachem(id):
    if zachem_id(id).count(id) >= 16:
        bot.send_message(id, zachem_message)
        print("send")
        zachem_id(id, 1)
        print("send")


# функция добавляет новых пользователей
def add_user(user_name, user_school_class):
    if get_user_info(user_name):
        zachem(user_name)
        cursor.execute('UPDATE userdata SET class=? WHERE usertelegramid=?', (user_school_class, user_name))
        conn.commit()
        return
    with conn:
        cursor.execute("INSERT INTO userdata (usertelegramid, class) VALUES (?, ?)", (user_name, user_school_class))
        print("User added successfully")


# проверка, есть ли пользователь в базе данных
def get_user_info(user_telegram_id: str):
    cursor.execute('SELECT * FROM userdata WHERE usertelegramid = ?', (user_telegram_id,))
    row = cursor.fetchone()
    if row is not None:
        return True
    else:
        return False


# получение класса пользователя
def get_class_from_id(user_id: str):
    cursor.execute('SELECT * FROM userdata WHERE usertelegramid =?', (user_id,))
    row = cursor.fetchone()
    if row is not None:
        return row[2]
    else:
        return 'error'


# get temp cpu (only for paspberry pi)
def get_cpu_temperature():
    res = os.popen('vcgencmd measure_temp').readline()
    return float(res.replace("temp=", "").replace("'C", ""))


# блок функций для сохранения переменных в файлы
def say(s=""):
    if len(s) != 0:
        f = open("say.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("say.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def timeprow(s=""):
    if len(s) != 0:
        f = open("time.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("time.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def nametab(s=""):
    if len(s) != 0:
        f = open("name.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("name.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def date(s=""):
    if len(s) != 0:
        f = open("date.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("date.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def date0(s=""):
    if len(s) != 0:
        f = open("date0.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("date0.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def date1(s=""):
    if len(s) != 0:
        f = open("date1.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("date1.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def date2(s=""):
    if len(s) != 0:
        f = open("date2.txt", "w")
        f.write(s)
        return None
    elif s == "":
        f = open("date2.txt", "r")
        y = f.readline()
        return ("" if y == "~%~$~~!" else y)
    else:
        return None


def done(s=""):
    if len(s) != 0:
        f = open("done.txt", "w")
        f.write(str(s))
        return None
    elif s == "":
        f = open("done.txt", "r")
        y = f.readline()
        return ("" if y == "qwe" else y)
    else:
        return None


def komy(s=""):
    if len(s) != 0:
        f = open("komy.txt", "w")
        f.write(str(s))
        return None
    elif s == "":
        f = open("komy.txt", "r")
        y = f.readline()
        return int("1" if y == "000" else y)
    else:
        return None


@log
def zachem_id(s=0, w=0):
    if s != 650093184:
        f = open("zachem_id.txt", "r+")
        y = f.readline()
        s = str(s)
        if w:
            f.close()
            f = open("zachem_id.txt", "w")
            y = list(map(int, y.split()))
            y = list(set(y))
            stt = ""
            for i in y:
                stt += str(i)
                stt += " "
            y = stt
            f.write(y)
            return 1
        if s != "0":
            f.write(" " + str(s))
        y = y.replace("  ", " ")
        y = list(map(int, y.split()))
        return ("1" if y == "000" else y)
    return [0]


def subscribers(s=""):
    if len(s) != 0:
        f = open("subscribers.txt", "r+")
        f.write(str(subscribers()) + " " + str(s))
        return None
    elif s == "":
        f = open("subscribers.txt", "r")
        y = f.readline()
        return ("1" if y == "000" else y)
    else:
        return None


def unsubscribers(id=""):
    s = str(subscribers())
    if len(id) != 0:
        f = open("subscribers.txt", "w")
        s = s.replace("  ", " ")
        x = s.replace(id, "")
        f.write(x)
        return None
    else:
        return "Error"


def admins(s=""):
    if len(s) != 0:
        f = open("admins.txt", "r+")
        f.write(str(admins()) + " " + str(s))
        return None
    elif s == "":
        f = open("admins.txt", "r")
        y = f.readline()
        return ("1" if y == "000" else str(y))
    else:
        return None


def unadmins(id=""):
    s = str(admins())
    if len(id) != 0:
        f = open("admins.txt", "w")
        s = s.replace("  ", " ")
        x = s.replace(id, "")
        f.write(x)
        return None
    else:
        return "Error"


@log
def sendtome(s):
    bot.send_message(650093184, f"{s}")
    return str(650093184) + f"{s}"


@log
def sendto(id, s):
    bot.send_message(id, f"{s}")
    return str(id) + f"{s}"


def version():
    s = ""
    list = [
        "Ver1.0 - релизная версия, с 15.04.23 по 15.05.23\nИмеет функции: получение расписания по буква_номеру класса, функцию привязки /буква_номер, /start, /help, /info, /getmyid, /rasp, /authors, а так же функции принудительной перезагрузки, обновления таблицы и вывода динмического предупреждения",
        "Ver1.1 - с 15.05.23 по 31.08.23\nНовые функции: /version, некоторая оптимизация кода, улучшение его читаемости, изменение wiwod на wiwod2 (пробелы перед кабинетом), упрощение возможной смены оборудвания, появления второго бота для тестов: https://t.me/sg130testbot, новая ф-ия prow2 для ускорения получения рассписания в 4 раза",
        "Ver2.0 - с 31.08.23 по сегодня\nПереезд бота с сетевого города на самообеспечение, утрата функций polling() и prow2() за ненадобостью.\nДобавленна возможность подписаться на расписание /subscribe и /unsubscribe, админы теперь в отдельном файле, изменение /info, добавленна возможность поддержать проект, можно назначать и снимать с должностей админов, возможность узнать кто админ, так же добавленна возможность смотреть не только последнее расписание, а по дням (сегодня/завтра).\nДобавленна mute\\unmute и remove для исправления ошибок вручную, так же добавлены авто-поздравления с праздниками\nА ещё у нас вчера вышел из строя сервер и мы потеряли бд :("]
    for i in list:
        s += i + "\n" + "\n"
    return s


# функция формирования сообщеня для команды help
def helptext(s):
    otvet = ""
    admin = list(map(int, str(admins()).split()))
    text = ["11Б - чтобы получить рассписание 11Б класса", "11Б+ - чтобы получить рассписание 11Б класса на завтра",
            "11-   - получить расписание всей паралели", "11-+   - получить расписание всей паралели на завтра",
            "/11Б - что бы привязать свой аккаунт к 11Б классу", "/subscribe - подписаться на обновление расписания",
            "/unsubscribe - отписаться от обновления расписания", "/info - получить общую инормацию",
            "/settings - для изменения настроек вывода", "/start - запустить бота, увидеть стартовые сообщения",
            "/help - вызвать подсказку (это сообщение)", "/getmyid - для получения своего айди чата",
            "/version - для получения списка версий", "/authors - что бы увидеть имена авторов проекта",
            "/donat - админу на печеньки (на починку сервера, что бы бот был онлайн 24/7)"]
    textforadm = [
        ">say< - вывести предупреждение, указанное после пробела, перед каждым сообщением от бота (say no для отмены вывода этих сообщений)",
        ">sudo reboot< - принудительно перезагрузить сервер, если он онлайн",
        ">whoisadmin< для получения списка администраторов",
        ">adm XXXXXXXXXX<, где все X заменить на айди пользавателя - что бы сделать его администратором",
        ">unadm XXXXXXXXXX<, где все X заменить на айди пользавателя - что бы разжаловать его из администраторов",
        ">remove X.xlsx< - где X.xlsx полное название файла который надо удалить из всех папок системы, использавать только при ошибках загрузки расписания",
        ">mute X< - для включения режима молчания (бот будет выдавать только сообщение заглушку с указанием причины X) X - причина включения режима молчания",
        ">unmute< - для отключения режима молчания",
        "🗓Для добавления таблицы расписания просто отправьте её боту с аккаунта, имеющего права адимнистратора. Таблица, которую вы отправляете, должна иметь название начинающиеся с даты в формате ДД.ММ.ГГГГ, дальше может быть указанно что угодно"]
    for i in range(len(text)):
        otvet += text[i]
        otvet += "\n"
    if s in admin:
        otvet += "***Всё указанное далее будет работать только в том случае, если вы будите писать это с аккаунта, имеющего права администратора (этого)***\n"
        for j in range(len(textforadm)):
            otvet += textforadm[j]
            otvet += "\n"
    return otvet


@log
def mailing(name=""):
    list = map(int, str(subscribers()).split())
    for i in list:
        # time.sleep(0.1)
        try:
            if len(say()) != 0: bot.send_message(i, say())  # type: ignore
            bot.send_message(i, str(wiwod2(xxx(get_class_from_id(f'{i}'), 2, i, 0, name), i, 2,
                                           get_class_from_id(f'{i}'))))
            setactive(i)
        except:
            logf.write(f"[{datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')}] EXCEPT IN mailing WITH {i}\n")
            continue
    return "norm status"


def mute(s="не указана"):
    if s == "внутренняя проверка":
        f = open("mute.txt", "r")
        y = f.readline()
        if len(y) < 5:
            return 0
        else:
            return 1
    else:
        if len(s) != 0:
            f = open("mute.txt", "w")
            f.write(str(s))
            return None
        elif s == "":
            f = open("mute.txt", "r")
            y = f.readline()
            return ("" if y == "del" else y)
        else:
            return None


def setactive(id):
    cursor.execute(f"UPDATE userdata SET active=1 WHERE usertelegramid=?", (id,))
    conn.commit()


def make_unactive(s=""):
    now = datetime.datetime.now()
    if now.strftime("%HH") == "00" or now.strftime("%HH") == 00:
        cursor.execute(f"UPDATE userdata SET active=0")
        conn.commit()
    if s != "":
        cursor.execute(f"UPDATE userdata SET active=0")
        conn.commit()


def ban(user_id):
    if not (int(user_id) in list(map(int, str(admins()).split()))) or not (user_id == 650093184):
        current_time_now = time.time()
        if user_id in spam_count:
            if spam_count[user_id]["count"] >= SPAM_LIMIT:
                if current_time_now - spam_count[user_id]["last_time"] < SPAM_TIME_LIMIT:
                    if int(user_id) in list(map(int, str(admins()).split())) or user_id == 650093184:
                        pass
                    else:
                        return 1
        # Добавление или обновление информации о количестве сообщений для текущего пользователя
        if user_id in spam_count:
            spam_count[user_id]["count"] += 1
            spam_count[user_id]["last_time"] = current_time_now
        else:
            spam_count[user_id] = {"count": 1, "last_time": current_time_now}


def vasiliev(name):
    global kyda
    f = open(f"{kyda}{name}.xlsx", "rb")
    bot.send_document(650093184, f)
    f.close
    f = open(f"{kyda}{name}.xlsx", "rb")
    bot.send_document(650093184, f)
    f.close


@log
def givexlsx(id, sz):
    if sz:
        szsz = "завтра"
    else:
        szsz = "сегодня"
    if sz:
        try:
            skip = 1
            if int(datetime.datetime.today().weekday()) == 4:
                skip = 3
            elif int(datetime.datetime.today().weekday()) == 5:
                skip = 2
            name = (datetime.date.today() + datetime.timedelta(days=skip)).strftime('%d.%m.%Y')
        except:
            return "У нас ошибка!"
    else:
        try:
            name = (datetime.date.today()).strftime('%d.%m.%Y')
        except:
            return "У нас ошибка!"
    try:
        get_files = os.listdir(kyda)
        if f"{name}.xlsx" in get_files:
            f = open(f"{kyda}{name}.xlsx", "rb")
            bot.send_document(id, f)
            f.close
        else:
            return f"У нас нет рассисания на {szsz}"
    except:
        return "У нас ошибка!"
    return "Успешно!"


@log
def who_is_izm():
    b = []
    s = []
    a = teacher + ['5А', '5Б', '5В', '5Г', '5Д', '5Е', '6А', '6Б', '6В', '6Г', '6Д', '6Е',
                   '7А', '7Б', '7В', '7Г', '7Д', '7Е', '8А', '8Б', '8В', '8Г', '8Д', '8Е', '9А', '9Б', '9В', '9Г', '9Д',
                   '9Е',
                   '10А', '10Б', '10В', '10Г', '10Д', '10Е', '11А', '11Б', '11В', '11Г', '11Д', '11Е']
    for i in a:
        try:
            x = xxx(i, 2, 1)
            time.sleep(0.1)
            y = xxx(i, 2, 1, 1)
            if x == y:
                continue
            else:
                b.append(i)
        except:
            continue
    for j in b:
        cursor.execute(f"SELECT usertelegramid FROM userdata WHERE class=?", (j,))
        s.append(cursor.fetchall())
    os.remove(f"{kyda}1.xlsx")
    g = [j[0] for i in s for j in i]
    try:
        bot.send_message(650093184, str(f"{b}"))
    except:
        pass
    try:
        bot.send_message(650093184, str(f"{g}"))
    except:
        pass
    return g


@log
def izm(x):
    y = list(map(int, subscribers().split()))
    for i in x:
        if int(i) in y:
            time.sleep(0.1)
            try:
                bot.send_message(i, "❗️Изменение расписания❗️\n" + str(
                    wiwod2(xxx(get_class_from_id(f'{i}'), 2, i), i, 2, get_class_from_id(f'{i}'))))
                setactive(i)
            except:
                continue
    return "OK"


def send_to_class(id, s=""):
    a = []
    cursor.execute(f"SELECT usertelegramid FROM userdata WHERE class=?", (id,))
    a.append(cursor.fetchall())
    y = list(map(int, subscribers().split()))
    g = [i[0] for i in a[0]]
    if s:
        st = s
    else:
        st = str(wiwod2(xxx(get_class_from_id(f'{i}'), 2, i), i, 2, get_class_from_id(f'{i}')))
    for i in g:
        if int(i) in y:
            time.sleep(0.1)
            try:
                bot.send_message(i, st)
                setactive(i)
            except:
                continue


def delete_lost():
    get_files = os.listdir(kyda)
    for i in range(5, 12):
        name_lost = (datetime.date.today() - datetime.timedelta(days=i)).strftime('%d.%m.%Y')
        if name_lost in get_files:
            os.remove(kyda + name_lost + ".xlsx")


# функция запоминания отправки слова "done" по id пользователя
def send_done(ID):
    bot.send_message(ID, 'done')


# обработчик файлов полученых ботом
@bot.message_handler(content_types=['document'])
def handle_docs(message):
    if mute("внутренняя проверка"):
        if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
            bot.send_message(message.chat.id,
                             f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
        else:
            bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
            return "mute"
    else:
        if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
            try:
                file_info = bot.get_file(message.document.file_id)
                downloaded_file = bot.download_file(file_info.file_path)
                src = otkyda + message.document.file_name
                with open(src, 'wb') as new_file:
                    new_file.write(downloaded_file)
                nametab(message.document.file_name)
                now = datetime.datetime.now()
                timeprow(now.strftime("%H:%M:%S"))
                bot.reply_to(message, "Принято")
                #  переименовываем файл

                f = f"{otkyda}{str(message.document.file_name)}"
                off = f"{kyda}{date0()}.xlsx"
                fff = f"{kyda}"
                flag_izm = 0
                get_files_2 = os.listdir(kyda)
                name = str(message.document.file_name)[:10]
                """"""
                if name[-1] == " ":
                    name = name[:-1]
                    if re.fullmatch(r'\d{2}', name[2:]) is None:
                        name = "0" + name
                """"""
                if f"{date0()}.xlsx" in get_files_2: os.remove(off)
                ff = f"{otkyda}{name}.xlsx"
                get_files1 = os.listdir(otkyda)
                if f"{name}.xlsx" in get_files1: os.remove(f"{otkyda}{name}.xlsx")
                os.rename(f, ff)
                if f"{name}.xlsx" in get_files_2:
                    flag_izm = 1
                    os.rename(f"{kyda}{name}.xlsx", f"{kyda}1.xlsx")
                shutil.move(ff, fff)
                try:
                    vasiliev(name)
                except:
                    bot.send_message(650093184, f"Ошибка при Васильеве")
                delete_lost()
                if flag_izm:
                    izm(who_is_izm())
                    print(f"done izm")
                    bot.reply_to(message, "Расписание успешно изменено")
                else:
                    date0(str(date1()))
                    date1(str(date2()))
                    date2(name)
                    try:
                        mailing(name)
                    except:
                        pass

            except Exception as e:
                bot.reply_to(message, e)  # type: ignore
                bot.send_message(650093184, f"Ошибка при загрузке расписания:\n{e}")  # type: ignore
        else:
            bot.send_message(650093184, str(message.from_user.id))
            bot.forward_message(650093184, message.chat.id, message.id)
            if message.chat.id > 0: bot.send_message(message.chat.id, "Вы не являетесь администратором")


# обработчик сообщений полученых ботом
@bot.message_handler(commands=['start'])
def start(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if message.chat.id < 0:
            bot.send_message(message.chat.id,
                             'Для групп работают все функции и вызов расписания расписания по ключевым сообщениям "/today" и "/tomorrow"')
        else:
            if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btn1 = types.KeyboardButton("На сегодня")
            btn2 = types.KeyboardButton("На завтра")
            markup.add(btn1, btn2)
            bot.send_message(message.chat.id,
                             text="Это тестовый бот для вывода рассписания с сайта sg, обрабатывает не более 5 сообщений в 20 секунд, больше информации - /help".format(
                                 message.from_user), reply_markup=std_keyboard)
            bot.send_message(message.chat.id,
                             "Внимание! Если вы видите что бот отправляет как-то странно, проверте расписание другим способом, бот не совершнен и может ошибиться")
            bot.send_message(message.chat.id, "Если вы увидите серьёзную ошибку - пишите сюда: https://t.me/gaming_raspredval")
            bot.send_message(message.chat.id,
                             "Добавляя бота в группу tg, нужно понимать, что будет работать только функция привязки и функция вызова расписания привязанного класса")


@bot.message_handler(commands=['help'])
def help(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        btn1 = types.KeyboardButton("На сегодня")
        btn2 = types.KeyboardButton("На завтра")
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, helptext(message.chat.id))


@bot.message_handler(commands=['today'])
def today(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        setactive(message.from_user.id)
        if len(say()) != 0: bot.send_message(message.from_user.id, say())  # type: ignore
        # проверяет, есть ли пользователь в базе данных
        if get_user_info(f'{message.chat.id}'):
            bot.send_message(message.chat.id,
                             str(wiwod2(xxx(get_class_from_id(f'{message.chat.id}'), 1, message.chat.id),
                                        message.chat.id, 1, get_class_from_id(f'{message.chat.id}'))))
        else:
            bot.send_message(message.chat.id,
                             "Вы еще не зарегистрированы, укажите свой класс текстом в чате с указаним слэша, например так: /11Б")


@bot.message_handler(commands=['tomorrow'])
def tomorrow(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        setactive(message.from_user.id)
        if len(say()) != 0: bot.send_message(message.from_user.id, say())  # type: ignore
        # проверяет, есть ли пользователь в базе данных
        if get_user_info(f'{message.chat.id}'):
            bot.send_message(message.chat.id,
                             str(wiwod2(xxx(get_class_from_id(f'{message.chat.id}'), 2, message.chat.id),
                                        message.chat.id, 2, get_class_from_id(f'{message.chat.id}'))))
        else:
            bot.send_message(message.chat.id,
                             "Вы еще не зарегистрированы, укажите свой класс текстом в чате с указаним слэша, например так: /11Б")


@bot.message_handler(commands=['settings'])
def settings(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"
        settin = str(get_settings_from_id(message.chat.id))
        settin = settin[settin.find("'") + 1:settin.rfind("'")]
        if message.chat.id < 0:
            bot.send_message(message.chat.id,
                             f'Ваши настройки сейчас: {settin}\n\nЧто бы указать настройки напишите /XXXXXXXX где каждый X отображает да(1)/нет(0) на вопрос далее (принимается только 8 символов, если вопросов меньше чем 8, то просто добавте в конце нули, так, что бы в итоге было 8 символов, "/" - обязателен!):\n\n1️⃣ Нужно ли отображать полные названия предметов без сокращений?\n\n2️⃣Отключить вывод даты (шапки сообщения)?\n\n3️⃣ Выводить ли класс, для которого предоставленно расписание?\n\n4️⃣Отказаться от поздравлений на праздники?\n\n5️⃣Выводить ли номер урока? (временно не работает для учителей, сделаем по запросу)\n\n6️⃣Сокращать ли вывод времени уроков? (временно не работает для учителей, сделаем по запросу)\n\n7️⃣Настройка для учителей: отключить ли вывод стандартного для вас урока? (информатику перестанет показывать "урок информатика", но вот все классные часы будет показывать как обычно, что бы к ним можно было подготовиться)')
        else:
            if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
            bot.send_message(message.chat.id,
                             f'Ваши настройки сейчас: {settin}\n\nЧто бы указать настройки напишите /XXXXXXXX где каждый X отображает да(1)/нет(0) на вопрос далее (принимается только 8 символов, если вопросов меньше чем 8, то просто добавте в конце нули, так, что бы в итоге было 8 символов, "/" - обязателен!):\n\n1️⃣ Нужно ли отображать полные названия предметов без сокращений?\n\n2️⃣Отключить вывод даты (шапки сообщения)?\n\n3️⃣ Выводить ли класс, для которого предоставленно расписание?\n\n4️⃣Отказаться от поздравлений на праздники?\n\n5️⃣Выводить ли номер урока? (временно не работает для учителей, сделаем по запросу)\n\n6️⃣Сокращать ли вывод времени уроков? (временно не работает для учителей, сделаем по запросу)\n\n7️⃣Настройка для учителей: отключить ли вывод стандартного для вас урока? (информатику перестанет показывать "урок информатика", но вот все классные часы будет показывать как обычно, что бы к ним можно было подготовиться)')


@bot.message_handler(commands=['info'])
def info(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"
        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        info = ""
        info += f"Зарег. пользователей: {count_users() - 2}\n"
        info += f"Пользователей активных за день: {active_users()}\n"
        info += f"Подключённых групп: {count_grupp() - 1}\n"
        info += f"Подписавшихся на рассылку: {len(list(map(int, str(subscribers()).split())))}\n"
        info += f"Админов: {len(list(map(int, str(admins()).split())))}\n"
        """Раскоментить на малинке"""
        # try: info += f"Температура сервера: {get_cpu_temperature()}C\n"
        # except: pass
        info += f"Время последнего перезапуска: {reboot_time}\n"
        # info += f"Время последнего обновления .xlsx: {timeprow()}\n"
        bot.send_message(message.chat.id, info)


@bot.message_handler(commands=['authors'])
def authors(message):  # type: ignore
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        bot.send_message(message.chat.id,
                         "Паравозик Томас - не ебу причем тут паравоз")


@bot.message_handler(commands=['getmyid'])
def getmyid(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        bot.send_message(message.chat.id, "Ваш айди чата: " + str(message.chat.id))
        bot.send_message(message.chat.id, "Ваш айди личный: " + str(message.from_user.id))


@bot.message_handler(commands=['rasp'])
def rasp(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        bot.send_message(message.chat.id,
                         str(wiwod2(xxx(get_class_from_id(f'{message.chat.id}'), 1, message.chat.id), message.chat.id,
                                    1, get_class_from_id(f'{message.chat.id}'))))


@bot.message_handler(commands=['subscribe'])
def subscribe(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"
        # bot.send_message(message.chat.id, "1")

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore

        if str(message.chat.id) in str(subscribers()):
            bot.send_message(message.chat.id, "Вы уже подписаны\nОтказаться - /unsubscribe")
        else:
            subscribers(f" {message.chat.id}")
            bot.send_message(message.chat.id,
                             "Успешно! Теперь вам будет приходить расписание сразу же, как только оно появится.\nЧто бы отказаться от рассылки напишите /unsubscribe")


@bot.message_handler(commands=['unsubscribe'])
def unsubscribe(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"
        # bot.send_message(message.chat.id, "1")

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore

        if str(message.chat.id) in str(subscribers()):
            bot.send_message(message.chat.id, "Успешно! Теперь вам не будет приходить расписание")
            unsubscribers(f" {message.chat.id}")
        else:
            bot.send_message(message.chat.id, "Успешно! Но вы и так были не подписаны")


@bot.message_handler(commands=['version'])
def versions(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
        bot.send_message(message.chat.id, version())


@bot.message_handler(commands=['donat'])
def donat(message):
    if ban(message.from_user.id):
        pass
    else:
        if mute("внутренняя проверка"):
            if message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184:
                bot.send_message(message.chat.id,
                                 f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
            else:
                bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                return "mute"

        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore

        bot.send_message(message.chat.id, "5469 1600 1663 8376 Сбербанк\nСпасибо за поддержку!")


@bot.message_handler(commands=['givemetoday'])
def givemetoday(message):
    bot.send_message(message.chat.id, givexlsx(message.chat.id, 0))


@bot.message_handler(commands=['givemetomorrow'])
def givemetomorrow(message):
    bot.send_message(message.chat.id, givexlsx(message.chat.id, 1))


@bot.message_handler(content_types=['text'])
def get__schedule(message):
    try:
        if ban(message.from_user.id):
            pass
        else:
            if mute("внутренняя проверка"):
                if (message.chat.id in list(map(int, str(admins()).split()))) or message.chat.id == 650093184:
                    bot.send_message(message.chat.id,
                                     f"Бот переведён в режим молчания по причине: {mute('')}\nСообщения получают только администраторы, для выхода из режима молчания >unmute<")
                else:
                    bot.send_message(message.chat.id, f"Бот переведён в режим молчания по причине: {mute('')}")
                    return "mute"
            abba = str(message.text)
            abba = abba.replace(" ", "")
            abba = abba.replace(".", "")
            abba = abba.replace("'", "")
            abba = abba.replace('"', '')
            abba = abba.replace(" ", "")
            abba = abba.upper()
            abba = abba.replace("A", "А")
            abba = abba.replace("B", "Б")
            abba = abba.replace("V", "В")
            abba = abba.replace("G", "Г")
            abba = abba.replace("D", "Д")
            abba = abba.replace("E", "Е")
            # обработка групп
            if message.chat.id < 0:
                if len(abba) == 9:
                    if abba.count("1") + abba.count("0") == 8 and "/" in abba:
                        if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
                        cursor.execute(f"UPDATE userdata SET settings=? WHERE usertelegramid=?",
                                       (str(message.text[1:]), message.chat.id))
                        conn.commit()
                        bot.send_message(message.chat.id,
                                         f"Принято, теперь ваши настройки: {get_settings_from_id(message.chat.id)[0][0]}")
                    else:
                        pass
                elif "/" in abba:
                    if abba in awe:
                        add_user(message.chat.id, abba[1:])
                        bot.send_message(message.chat.id,
                                         text='Успешно, вы можете изменить это в любой момент'.format(
                                             message.from_user))
                return 0

            # Обработка сообщения в обычном режиме для обычных пользователей
            if "unmute" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                mute("del")
                bot.send_message(message.chat.id, "Бот выведен из режима молчания, сообщения теперь получают все")
            elif "mute" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                x = message.text[5:]
                mute(f"{x}")
                bot.send_message(message.chat.id,
                                 "Бот переведён в режим молчания, сообщения могут получать только администраторы")
            elif message.text == 'На сегодня':
                setactive(message.from_user.id)
                if len(say()) != 0: bot.send_message(message.from_user.id, say())  # type: ignore
                # проверяет, есть ли пользователь в базе данных
                if get_user_info(f'{message.chat.id}'):
                    bot.send_message(message.chat.id,
                                     str(wiwod2(xxx(get_class_from_id(f'{message.chat.id}'), 1, message.chat.id),
                                                message.chat.id, 1, get_class_from_id(f'{message.chat.id}'))),
                                     reply_markup=std_keyboard)
                else:
                    bot.send_message(message.chat.id,
                                     "Вы еще не зарегистрированы, укажите свой класс текстом в чате с указаним слэша, например так: /11Б")
                    bot.send_message(message.chat.id,
                                     "Или можете просто написать 11Б и получить рассписание этого класса на сегодня")
            elif message.text == 'На завтра':

                setactive(message.from_user.id)
                if len(say()) != 0: bot.send_message(message.from_user.id, say())  # type: ignore
                # проверяет, есть ли пользователь в базе данных
                if get_user_info(f'{message.chat.id}'):
                    bot.send_message(message.chat.id,
                                     str(wiwod2(xxx(get_class_from_id(f'{message.chat.id}'), 2, message.chat.id),
                                                message.chat.id, 2, get_class_from_id(f'{message.chat.id}'))),
                                     reply_markup=std_keyboard)

                else:
                    bot.send_message(message.chat.id,
                                     "Вы еще не зарегистрированы, укажите свой класс текстом в чате с указаним слэша, например так: /11Б")
                    bot.send_message(message.chat.id,
                                     "Или можете просто написать 11Б+ и получить рассписание этого класса на завтра")

            elif ': ' in message.text or 'Назад' in message.text or 'Настройки' in message.text:
                cur_settings = cursor.execute('SELECT settings FROM userdata WHERE usertelegramid=?',
                                              (message.chat.id,)).fetchone()[
                    0]  # Получаем настройки пользователя по его id

                if len(cur_settings) < 8:
                    cursor.execute('UPDATE userdata SET settings=? WHERE usertelegramid=?',
                                   (cur_settings + '0' * (8 - len(cur_settings)), message.chat.id))
                    conn.commit
                    cur_settings = cursor.execute('SELECT settings FROM userdata WHERE usertelegramid=?',
                                                  (message.chat.id,)).fetchone()[0]

                if cur_settings != None:
                    ask1, ask2, ask3, ask4, ask5, ask6, ask7 = ['нет' if int(i) else 'да' for i in
                                                                cur_settings[
                                                                :7]]  # Переменные -- ответы на вопросы, получаем из настроек

                    choose_sett = types.ReplyKeyboardMarkup(True, True)

                    der = {
                        '1: ': ask1,
                        '2: ': ask2,
                        '3: ': ask3,
                        '4: ': ask4,
                        '5: ': ask5,
                        '6: ': ask6,
                        '7: ': ask7
                    }  # В этом словаре содержатся: номер вопроса и ответ на него

                    ask1, ask2, ask3, ask4, ask5, ask6, ask7 = list(
                        der.values())  # Переопределяем значения ответов на вопросы, мало ли что
                    # Создаём 6 кнопок со значениями "Номер вопроса: ответ на него" + Назад
                    choose_sett.row(f'1: {ask1}', f'2: {ask2}', f'3: {ask3}', f'4: {ask4}')
                    choose_sett.row(f'5: {ask5}', f'6: {ask6}', f'7: {ask7}', 'Назад')
                    if message.text == 'Назад':
                        bot.send_message(message.chat.id, "Вы вышили из режима изменения настроек",
                                         reply_markup=std_keyboard)
                    if message.text == 'Настройки':
                        bot.send_message(message.chat.id,
                                         f'📎Сейчас ваши настройки - {cur_settings}\n\n✏️Для их изменения ответьте на вопросы ниже\n\n1️⃣ Нужно ли отображать полные названия предметов без сокращений?\n\n2️⃣Отключить вывод даты (шапки сообщения)?\n\n3️⃣ Выводить ли класс, для которого предоставленно расписание?\n\n4️⃣Отказаться от поздравлений на праздники?\n\n5️⃣Выводить ли номер урока? (временно не работает для учителей, сделаем по запросу)\n\n6️⃣Сокращать ли вывод времени уроков? (временно не работает для учителей, сделаем по запросу)\n\n7️⃣Настройка для учителей: отключить ли вывод стандартного для вас урока? (информатику перестанет показывать "урок информатика", но вот все классные часы будет показывать как обычно, что бы к ним можно было подготовиться)',
                                         reply_markup=choose_sett)

                    if ': ' in message.text:  # Если текст содержит ": ", то мы обращаемся к der и изменяем его соответствующее значение следующим образом.
                        if der[message.text[
                               :3]] == 'да':  # message.text[:3] - это номер вопроса, if der[message.text[:3]] - ответ на него.
                            der[message.text[:3]] = "нет"
                        else:
                            der[message.text[:3]] = "да"

                        cur_settings = ''.join(
                            ['0' if i == 'да' else '1' for i in list(der.values())])  # Пересобираем настройки из der

                        cursor.execute('UPDATE userdata SET settings=? WHERE usertelegramid=?',
                                       (cur_settings, message.chat.id))  # Записываем настройки в бд
                        conn.commit()

                        ask1, ask2, ask3, ask4, ask5, ask6, ask7 = list(
                            der.values())  # Переопределяем значения ответов на вопрос, на всякий случай

                        # Переобъявляем кнопки выбора настроек
                        choose_sett = types.ReplyKeyboardMarkup(True, True)
                        choose_sett.row(f'1: {ask1}', f'2: {ask2}', f'3: {ask3}', f'4: {ask4}')
                        choose_sett.row(f'5: {ask5}', f'6: {ask6}', f'7: {ask7}', 'Назад')

                        bot.send_message(message.chat.id, f'Принято, ваши настройки - {cur_settings}',
                                         reply_markup=choose_sett)  # Ну тут ты и сам поймешь
                else:
                    bot.send_message(message.chat.id,
                                     "Вы ещё не зарегестрированны и этот функционал вам не доступен\nЗарегестрируйтесь, написав сначала /, а затем цифру и букву вашего класса, например: /11Б",
                                     reply_markup=std_keyboard)

            elif 'remove' in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                x = message.text[7:]
                try:
                    get_files1 = os.listdir(otkyda)
                    if f"{x}" in get_files1: os.remove(f"{otkyda}{x}")
                    get_files2 = os.listdir(kyda)
                    if f"{x}" in get_files2: os.remove(f"{kyda}{x}")
                    bot.reply_to(message, "Успешно!")
                except Exception as e:
                    bot.reply_to(message, e)  # type: ignore
            elif "whoisadmin" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                global im_done
                user_id = message.from_user.id
                current_time_now = time.time()
                # daF = str(message.text); daF = daF.replace("/", "")
                if user_id in spam_count:
                    if spam_count[user_id]["count"] >= SPAM_LIMIT:
                        if current_time_now - spam_count[user_id]["last_time"] < SPAM_TIME_LIMIT:
                            return
                # Добавление или обновление информации о количестве сообщений для текущего пользователя
                if user_id in spam_count:
                    spam_count[user_id]["count"] += 1
                    spam_count[user_id]["last_time"] = current_time_now
                else:
                    spam_count[user_id] = {"count": 1, "last_time": current_time_now}

                if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore

                bot.send_message(message.chat.id, str(list(map(int, str(admins()).split()))))
            elif "unadm" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                x = message.text[6:]

                if len(x) > 6:
                    if not (int(x) in list(map(int, str(admins()).split()))):
                        bot.send_message(message.chat.id, f"Готово, но {x} и так не был адимном")
                    else:
                        try:
                            unadmins(x)
                            bot.send_message(message.chat.id,
                                             f"Успешно! Пользователь {x} больше не является администратором")
                        except:
                            bot.send_message(message.chat.id,
                                             "Непредвиденная ошибка, обратитесь к разработчику или попробуйте позже")
                else:
                    bot.send_message(message.chat.id,
                                     "Не верный формат, укажите в формате unadm XXXXXXXXXX, где все X заменить на айди пользователя которого вы намерены лишить прав адиминистратора\nДля получения айди пользавателя воспользуётесь с его аккаунта функцией /getmyid\nДля получения списка всех администраторов воспользуйтесь >whoisadm<")
            elif "adm" in message.text and not ("un" in message.text) and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                x = message.text[4:]
                if len(x) > 6:
                    if int(x) in list(map(int, str(admins()).split())):
                        bot.send_message(message.chat.id, "Этот пользаватель и так уже админ, куда выше?")
                    else:
                        try:
                            admins(x)
                            bot.send_message(message.chat.id,
                                             f"Успешно! Пользователю с айди {x} присвоен статус администратора")
                            bot.send_message(x, "Теперь вы администратор")
                        except:
                            bot.send_message(message.chat.id,
                                             "Непредвиденная ошибка, обратитесь к разработчику или попробуйте позже")
                else:
                    bot.send_message(message.chat.id,
                                     "Не верный формат, укажите в формате adm XXXXXXXXXX, где все X заменить на айди пользователя которого вы намерены сделать адиминистратором\nДля получения айди пользавателя воспользуётесь с его аккаунта функцией /getmyid")
            elif message.text == 'sudo reboot' and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                bot.send_message(message.chat.id, "please wait ~4 min")
                komy(f"{message.chat.id}")
                time.sleep(2)
                reboot()
            elif (message.text.isdigit() or message.text in ['Геол.музей', 'Б.спорт.зал', 'Ист.музей']) and (
                    message.text.count("1") + message.text.count("0") != 8):
                bot.send_message(message.chat.id,
                                 str(wiwod2(xxx(message.text, 1, message.chat.id), message.chat.id, 1, message.text)))
            elif abba in qwq and "-" in abba:
                if not ("+" in abba):
                    day = 1
                else:
                    day = 2
                    abba = abba[:-1]
                setactive(message.from_user.id)
                if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
                text = ""
                if day == 1:
                    data = str(wiwod2(xxx((abba[:abba.find("-")] + "А"), day, 1), 1, 1, (abba[:abba.find("-")] + "А")))
                else:
                    data = str(wiwod2(xxx((abba[:abba.find("-")] + "А"), day, 1), 1, 2, (abba[:abba.find("-")] + "А")))
                try:
                    data = data[data.find("🗓"):(data.find(":") - 4)]
                except:
                    data = "Не известна"
                text = text + data + "\n\n"
                try:
                    for i in "АБВГДЕ":
                        text_to_append = str(
                            wiwod2(xxx((abba[:abba.find("-")] + i), day, 0), 0, 1, (abba[:abba.find("-")] + i))) + "\n"
                        if len(text_to_append) > 30:
                            text = text + text_to_append
                except:
                    pass
                bot.send_message(message.chat.id, text)
            elif "/" in abba:
                if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
                if abba in awe:
                    print(abba, message.chat.id)
                    add_user(message.chat.id, abba[1:])
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    btn1 = types.KeyboardButton("На сегодня")
                    btn2 = types.KeyboardButton("На завтра")
                    markup.add(btn1, btn2)
                    bot.send_message(message.chat.id,
                                     text='Успешно, вы можете изменить это в любой момент'.format(message.from_user),
                                     reply_markup=std_keyboard)
                    # cursor.execute(f"UPDATE userdate SET settings=? WHERE id=?",(10000000,650093184))
            elif abba in qwq:
                if not ("+" in abba):
                    day = 1
                else:
                    day = 2
                    abba = abba[:-1]
                setactive(message.from_user.id)
                if len(say()) != 0: bot.send_message(message.chat.id, say())  # type: ignore
                bot.send_message(message.chat.id,
                                 str(wiwod2(xxx(abba, day, message.chat.id), message.chat.id, 1, abba)))
            elif "say" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                if "no" in message.text:
                    say("~%~$~~!")
                else:
                    say(message.text[3:])
                bot.send_message(message.chat.id, "accepted")
            elif "eval" in message.text and (
                    message.chat.id in list(map(int, str(admins()).split())) or message.chat.id == 650093184):
                x = message.text[5:]
                eval(f"{x}")
                bot.send_message(message.chat.id, "done")
            elif len(abba) == 8:
                if abba.count("1") + abba.count("0") == 8:
                    if len(say()) != 0: bot.send_message(message.from_user.id, say())  # type: ignore
                    cursor.execute(f"UPDATE userdata SET settings=? WHERE usertelegramid=?",
                                   (str(message.text), message.chat.id))
                    conn.commit()
                    bot.send_message(message.from_user.id,
                                     f"Принято, теперь ваши настройки: {get_settings_from_id(message.chat.id)[0][0]}")
                else:
                    pass  # bot.send_message(message.from_user.id, "Ошибка синтаксиса")
            else:
                if len(message.chat.id) > 0:
                    if len(say()) != 0: pass  # bot.send_message(message.from_user.id, say()) # type: ignore
                    # bot.send_message(message.from_user.id, "Ошибка синтаксиса")
    except Exception as e:
        pass  # bot.send_message(650093184, e)  # type: ignore



# По приколу
@bot.message_handler(content_types=['photo', 'text'], func=lambda message: True)
def handle_photo(message):
    if message.photo:
        photo = message.photo[-1]
        file_id = photo.file_id
        file_info = bot.get_file(file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        bot.send_message(650093184, str(message.from_user.id))
        bot.send_photo(650093184, downloaded_file, caption=message.caption)



def botara():
    while True:
        try:
            bot.polling(none_stop=True, interval=0)
            logf.write(f"[{datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')}] BOT STARTED POLLONG\n")
        except Exception as e:
            time.sleep(10)
            logf.write(f"[{datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')}] ERROR at botara with {e}\n")
            print(e)
            print("restart")
            continue


def prow():
    a = date()
    b = (datetime.date.today()).strftime('%d.%m.%Y')
    if not (a == b):
        make_unactive("1")
        date(b)


def nolnol():
    prow()
    time.sleep(3600)


# уведомление в консоль о запуске
print("i'am has been started")
logf.write(f"[{datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S')}] bot has been started\n")
# make_unactive("NOW")

# print(str(get_settings_from_id(5967500116)))
# отпрвка "done" тому, кто перезапустил систему принудительно
if komy() != 1:
    send_done(komy())
    komy("000")
# запуск обработки ботом трёх потоков
th1 = threading.Thread(target=botara)
#th2 = threading.Thread(target=nolnol)
# th2 = threading.Thread(target=planreboot)
#cursor.execute(f"DELETE FROM userdata;")
th1.start()
#th2.start()
# вызовы функций для проверок
#cursor.execute(f"UPDATE userdata SET settings=? WHERE id=?",(00000000,650093184))
